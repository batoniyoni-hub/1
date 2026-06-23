# -*- coding: utf-8 -*-
import os
import datetime
import pyotp
import json
import sqlite3
import pickle
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from account_database import AccountDatabase, AccountStatus
from typing import Dict, List, Optional
import threading
from concurrent.futures import ThreadPoolExecutor
import requests

class MultiAccountManager:
    """Менеджер мультиаккаунтов с БД и поддержкой кампаний"""
    
    def __init__(self, geo, count, campaign_id=None):
        self.geo = geo
        self.count = count
        self.campaign_id = campaign_id
        self.date_str = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.filename = f"Accounts_{geo}_{count}acc_{self.date_str}.xlsx"
        
        # Инициализация БД
        self.db = AccountDatabase()
        self.wb = Workbook()
        self.setup_excel()
        self._lock = threading.Lock()
        
        # Статистика
        self.stats = {
            'created': 0,
            'valid': 0,
            'failed': 0,
            'checkpointed': 0
        }
    
    def setup_excel(self):
        """Создание Excel файла с красивым оформлением"""
        ws = self.wb.active
        ws.title = "Accounts Data"
        
        headers = [
            'ID БД', 'Платформа', 'Статус', 'Username', 'Email', 'Email Pass',
            'Account Pass', '2FA Secret', 'Token', 'API Key',
            'Имя', 'Фамилия', 'ДР', 'Пол', 'Страна', 'Город',
            'Работа', 'Компания', 'ВУЗ', 'Школа',
            'Прокси', 'Device ID', 'Trust Score', 'Друзей', 'Фолловеров',
            'BM ID', 'Ad Account', 'BM Token', 'Статус БМ',
            'Дата создания', 'Последний вход', 'Куки', 'Примечания'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Установка ширины колонок
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        self.wb.save(self.filename)
    
    def save_account_to_db(self, account_data: Dict) -> Optional[int]:
        """Сохранить аккаунт в БД"""
        try:
            # Подготовка данных
            db_data = {
                'platform': account_data.get('platform', 'facebook'),
                'account_id': account_data.get('account_id'),
                'username': account_data.get('username'),
                'email': account_data.get('email'),
                'email_password': account_data.get('email_password'),
                'account_password': account_data.get('account_password'),
                'first_name': account_data.get('first_name'),
                'last_name': account_data.get('last_name'),
                'phone': account_data.get('phone'),
                'birthday': account_data.get('birthday'),
                'gender': account_data.get('gender'),
                'country': account_data.get('country', self.geo),
                'city': account_data.get('city'),
                'profile_pic_url': account_data.get('profile_pic_url'),
                'profile_bio': account_data.get('profile_bio'),
                'job_title': account_data.get('job_title'),
                'job_company': account_data.get('job_company'),
                'education_school': account_data.get('education_school'),
                'education_university': account_data.get('education_university'),
                'access_token': account_data.get('access_token'),
                'refresh_token': account_data.get('refresh_token'),
                'api_key': account_data.get('api_key'),
                'two_fa_secret': account_data.get('two_fa_secret'),
                'two_fa_enabled': 1 if account_data.get('two_fa_secret') else 0,
                'cookies': account_data.get('cookies'),
                'session_data': account_data.get('session_data'),
                'user_agent': account_data.get('user_agent'),
                'proxy_ip': account_data.get('proxy'),
                'antidetect_fingerprint': account_data.get('antidetect_fingerprint'),
                'device_id': account_data.get('device_id'),
                'android_id': account_data.get('android_id'),
                'status': account_data.get('status', 'valid'),
                'trust_score': account_data.get('trust_score', 0.0),
                'business_manager_id': account_data.get('bm_id'),
                'ad_account_id': account_data.get('ad_id'),
                'bm_token': account_data.get('bm_token'),
                'business_email': account_data.get('biz_email'),
                'notes': account_data.get('notes'),
            }
            
            # Фильтруем None значения
            db_data = {k: v for k, v in db_data.items() if v is not None}
            
            account_id = self.db.add_account(db_data)
            
            if account_id and self.campaign_id:
                self.db.assign_accounts_to_campaign(self.campaign_id, [account_id])
            
            return account_id
        except Exception as e:
            print(f"[-] Ошибка сохранения в БД: {e}")
        return None
    
    def add_account_to_excel(self, data: Dict, db_id: int = None):
        """Добавить аккаунт в Excel"""
        try:
            with self._lock:
                ws = self.wb.active
                next_row = ws.max_row + 1
                
                # Генерация 2FA кода
                two_fa_code = ""
                if data.get('2fa_secret'):
                    try:
                        totp = pyotp.TOTP(data['2fa_secret'])
                        two_fa_code = totp.now()
                    except:
                        pass
                
                row_data = [
                    db_id or '',
                    data.get('platform', 'facebook'),
                    data.get('status', 'valid'),
                    data.get('username', ''),
                    data.get('email', ''),
                    data.get('email_password', ''),
                    data.get('account_password', ''),
                    data.get('2fa_secret', ''),
                    data.get('access_token', '')[:50] + '...' if data.get('access_token') else '',
                    data.get('api_key', ''),
                    data.get('first_name', ''),
                    data.get('last_name', ''),
                    data.get('birthday', ''),
                    data.get('gender', ''),
                    data.get('country', self.geo),
                    data.get('city', ''),
                    data.get('job_title', ''),
                    data.get('job_company', ''),
                    data.get('education_university', ''),
                    data.get('education_school', ''),
                    data.get('proxy', ''),
                    data.get('device_id', ''),
                    data.get('trust_score', 0.0),
                    data.get('friend_count', 0),
                    data.get('follower_count', 0),
                    data.get('bm_id', ''),
                    data.get('ad_id', ''),
                    data.get('bm_token', '')[:30] + '...' if data.get('bm_token') else '',
                    data.get('checkpoint', 'None'),
                    data.get('created_at', datetime.datetime.now().strftime('%Y-%m-%d %H:%M')),
                    data.get('last_login', ''),
                    'Сохранены' if data.get('cookies') else 'Нет',
                    data.get('notes', '')
                ]
                
                ws.append(row_data)
                
                # Раскрашивание строки по статусу
                if data.get('status') == 'valid':
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif data.get('status') == 'checkpointed':
                    fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif data.get('status') == 'banned':
                    fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    fill = None
                
                if fill:
                    for cell in ws[next_row]:
                        cell.fill = fill
                
                self.wb.save(self.filename)
                return True
        except Exception as e:
            print(f"[-] Ошибка добавления в Excel: {e}")
        return False
    
    def add_account(self, data: Dict) -> bool:
        """Добавить аккаунт в БД и Excel"""
        try:
            # Сохранить в БД
            db_id = self.save_account_to_db(data)
            
            # Добавить в Excel
            if self.add_account_to_excel(data, db_id):
                with self._lock:
                    if data.get('status') == 'valid':
                        self.stats['valid'] += 1
                    self.stats['created'] += 1
                return True
        except Exception as e:
            print(f"[-] Ошибка добавления аккаунта: {e}")
            self.stats['failed'] += 1
        return False
    
    def generate_2fa(self) -> str:
        """Генерировать 2FA секрет"""
        return pyotp.random_base32()
    
    def get_2fa_code(self, secret: str) -> str:
        """Получить текущий 2FA код"""
        try:
            totp = pyotp.TOTP(secret)
            return totp.now()
        except:
            return ""
    
    def save_session_cookies(self, account_id: int, cookies: Dict, session_data: Dict = None):
        """Сохранить куки сессии"""
        try:
            cookies_json = json.dumps(cookies)
            session_json = json.dumps(session_data) if session_data else None
            
            self.db.update_account(account_id, {
                'cookies': cookies_json,
                'session_data': session_json,
                'last_login': datetime.datetime.now().isoformat()
            })
        except Exception as e:
            print(f"[-] Ошибка сохранения куки: {e}")
    
    def load_session_cookies(self, account_id: int) -> Dict:
        """Загрузить куки сессии"""
        try:
            account = self.db.get_account(account_id)
            if account and account.get('cookies'):
                return json.loads(account['cookies'])
        except Exception as e:
            print(f"[-] Ошибка загрузки куки: {e}")
        return {}
    
    def farm_account(self, account_id: int, platform: str, token: str, proxy: str = None) -> bool:
        """Имитация фарминга аккаунта"""
        try:
            print(f"[*] Фарминг аккаунта {account_id} на {platform}...")
            
            # Добавить действие в БД
            action_data = {
                'action': 'farming',
                'platform': platform,
                'timestamp': datetime.datetime.now().isoformat()
            }
            self.db.add_action(account_id, 'farming', action_data)
            
            # Обновить статус
            self.db.update_account(account_id, {
                'status': AccountStatus.FARMING.value,
                'last_action': datetime.datetime.now().isoformat()
            })
            
            return True
        except Exception as e:
            print(f"[-] Ошибка фарминга: {e}")
        return False
    
    def solve_checkpoint(self, account_id: int, checkpoint_type: str) -> bool:
        """Попытка решить чекпоинт"""
        try:
            account = self.db.get_account(account_id)
            if not account:
                return False
            
            print(f"[*] Решение чекпоинта {checkpoint_type} для {account['username']}...")
            
            # Логирование чекпоинта
            self.db.update_account(account_id, {
                'checkpoint_type': checkpoint_type,
                'status': AccountStatus.CHECKPOINTED.value
            })
            
            return True
        except Exception as e:
            print(f"[-] Ошибка решения чекпоинта: {e}")
        return False
    
    def create_batch_accounts(self, count: int, platform: str = 'facebook') -> List[int]:
        """Создать батч аккаунтов"""
        account_ids = []
        try:
            for i in range(count):
                account_data = {
                    'platform': platform,
                    'status': 'pending'
                }
                db_id = self.save_account_to_db(account_data)
                if db_id:
                    account_ids.append(db_id)
            
            print(f"[+] Создано {len(account_ids)} аккаунтов в БД")
        except Exception as e:
            print(f"[-] Ошибка создания батча: {e}")
        
        return account_ids
    
    def get_stats(self) -> Dict:
        """Получить статистику"""
        return {
            'local_stats': self.stats,
            'db_stats': self.db.get_stats(),
            'filename': self.filename
        }
    
    def link_business_manager(self, account_id: int, bm_id: str, ad_id: str, 
                             bm_token: str, biz_email: str = None) -> bool:
        """Привязать Business Manager"""
        try:
            self.db.update_account(account_id, {
                'business_manager_id': bm_id,
                'ad_account_id': ad_id,
                'bm_token': bm_token,
                'business_email': biz_email
            })
            print(f"[+] BM привязан к аккаунту {account_id}")
            return True
        except Exception as e:
            print(f"[-] Ошибка привязки BM: {e}")
        return False
    
    def export_valid_accounts(self, output_file: str = None) -> str:
        """Экспортировать валидные аккаунты"""
        if not output_file:
            output_file = f"valid_accounts_{self.date_str}.txt"
        
        try:
            accounts = self.db.get_accounts_by_status('valid')
            with open(output_file, 'w', encoding='utf-8') as f:
                for account in accounts:
                    f.write(f"{account['email']}:{account['account_password']}\n")
            
            print(f"[+] Экспортировано {len(accounts)} аккаунтов в {output_file}")
            return output_file
        except Exception as e:
            print(f"[-] Ошибка экспорта: {e}")
        return None
    
    def get_export_formats(self) -> Dict[str, str]:
        """Получить доступные форматы экспорта"""
        return {
            'xlsx': self.filename,
            'json': f"accounts_{self.date_str}.json",
            'txt': f"accounts_{self.date_str}.txt",
            'csv': f"accounts_{self.date_str}.csv"
        }
